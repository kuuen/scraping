import time
from selenium import webdriver
import os
import shutil
import openpyxl

from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import ElementNotInteractableException
from selenium.common.exceptions import StaleElementReferenceException
from logging import getLogger, StreamHandler, DEBUG
logger = getLogger(__name__)
handler = StreamHandler()
handler.setLevel(DEBUG)
logger.setLevel(DEBUG)
logger.addHandler(handler)
logger.propagate = False


# ★導入方法★
# seleniumをインストール
#   pip install selenium
#
# https://sites.google.com/chromium.org/driver/downloads?authuser=0 にて使用しているchromeのバージョンのwebdriverをダウンロードする
# ダウンロードしたexeファイルをpython.exeと同じディレクトリに保存する（場所は環境変数で確認できる）
# 店舗情報.xlsxを実行ファイルの同ディレクトリに置く
# main() の引数に商品名を指定する

# WebDriverのインスタンスを作成
driver = webdriver.Chrome() 

# 対象ページを開く
def get(url) :
  driver.get(url)
  time.sleep(2) # 2秒待機

# Yahooショッピングを開いて商品を検索する
def main(s):
  ''' s:検索キーワード
  '''

  # 開く
  get('https://www.amazon.co.jp/')

  # 検索欄にキーワードを入力
  driver.find_element_by_id("twotabsearchtextbox").send_keys(s)

  time.sleep(2) # 2秒待機
  # 検索ボタンクリック
  driver.find_element_by_id("nav-search-submit-text").click()

  index = 0
  
  while True :
    linkstrs = driver.find_elements_by_css_selector('.a-size-base-plus.a-color-base.a-text-normal')

    # メインループ
    listLoop(linkstrs)

    # 次ボタン取得 pagenateの種類が複数ある両方確認する
    nextLinks = driver.find_elements_by_class_name('a-last')
    nextLinks2 = driver.find_elements_by_css_selector('.s-pagination-item.s-pagination-next.s-pagination-button.s-pagination-separator')

    # 次ボタンが無い場合終了
    if len(nextLinks) == 0 and len(nextLinks2) == 0:
      break

    # 終了条件 paginate の次のボタンが非活性の場合終了
    if len(driver.find_elements_by_css_selector('.a-disabled.a-last')) > 0 or \
        len(driver.find_elements_by_css_selector('.s-pagination-item.s-pagination-next.s-pagination-disabled')) > 0:
      break

    # 次のページへ
    for link in nextLinks:
      if '次へ' in link.text :
        link.click()
        time.sleep(5)
        break
    
    if len(nextLinks2) > 0:
      nextLinks2[0].click()
      time.sleep(5)

  logger.info('処理終了')
  driver.quit() # ブラウザを閉じる

def listLoop(linkstrs):
  """
  linkstrs  :対象のリンク
  """

  linkRireki = {}

  # リンク一覧を参照
  for linkstr in linkstrs :
    
    # リンクが空白のものがあった
    if linkstr.text == '':
      continue

    # リンクの文字列でリンクを特定
    links = {}
    try :
      links = driver.find_elements_by_partial_link_text(linkstr.text)
    except StaleElementReferenceException:
      # 失敗した場合再処理を行う
      time.sleep(5)
      links = driver.find_elements_by_partial_link_text(linkstr.text)

    # リンクが複数ある場合は工夫が必要
    if len(links) == 1 :
      link = links[0]
      linkRireki[linkstr.text] = 0
    else:
      # 対象のリンクを指定する
      if  linkRireki.get(linkstr.text) == None:
        linkRireki[linkstr.text] = 0
      else:
        linkRireki[linkstr.text] += 1

      link = links[linkRireki[linkstr.text]]
      
    try:
      # リンクを開く
      link.click()
    except ElementClickInterceptedException :
      # 同一商品が複数ある場合にexceptionを吐く場合がある
      logger.error('同一商品でのエラー')
      continue
    except ElementNotInteractableException :
      # レンダリング作成中に操作をしたためエラーしばらく待つと操作可能になる
      time.sleep()
      # リンクを開く　再トライ
      link.click()

    time.sleep(2) # 2秒待機

    # 新しいタブに切り替える
    driver.switch_to.window(driver.window_handles[1])

    # 業者ページを参照
    referGyousya()

    # 新しいタブを閉じる
    driver.close()

    # 閉じたタブから前のタブを切り替える
    driver.switch_to.window(driver.window_handles[0])

# 業者ページ処理
def referGyousya() :

  link = None
  list = None

  # 商品ページによって販売元リンクが異なる
  link = driver.find_elements_by_css_selector(".tabular-buybox-text.a-spacing-none")[1]

  if link.text == '' :
    # 値が空白の場合リンクを取得できていない
  
    # 通常の注文から要素を取得、クリックして展開
    driver.find_elements_by_css_selector(".a-column.a-span12.a-text-left.truncate")[1].click()
    time.sleep(2) # 2秒待機

  try :
    # id指定のリンクの場合はそれを使用する
    link = driver.find_element_by_id('sellerProfileTriggerId')
  except NoSuchElementException:
    logger.info('') # 何もしない

  # 販売元がamazonの場合は対象外
  if link.text == 'Amazon.co.jp':
    return

  try:
    # 販売元をクリック
    link.click()
  except ElementNotInteractableException :
    # レンダリング作成中に操作をしたためエラーしばらく待つと操作可能になる
    time.sleep(3)
    link.click()

  time.sleep(2) # 2秒待機

  list = getData()

  if list == None:
    return

  # エクセル書き込み
  witeExcel(list)

def getData() :
  zyouhous = driver.find_elements_by_css_selector('.a-unordered-list.a-nostyle.a-vertical')
  list = {}

  zyouhou = zyouhous[0].text.split('\n')

  # 住所以外
  for z in zyouhou :
    if z.find('販売業者') > -1:
      list['companyName'] = z.split(':')[1]
    elif z.find('運営責任者') > -1:
      list['operationManager'] = z.split(':')[1]
    elif z.find('店舗名') > -1:
      list['shopName'] = z.split(':')[1]
    elif z.find('電話番号') > -1:
      list['tel'] = z.split(':')[1]

  # 住所
  zyouhou = zyouhous[1]
  adress = zyouhou.text.split('\n')

  # 住所の配列数は可変 
  # 例　マンション名９９９:区名:鎌倉市:神奈川県:2470071:JP
  a = ''
  for i in range(len(adress) - 3, -1, -1):
    a += adress[i]

  list['adress1'] = a
  list['postCode'] = adress[len(adress) -2]
  
  # 沖縄以外の場合は対象外
  if list['adress1'].find('沖縄') == -1:
    return None
  else:
    return list

def kaisyaExist(companyName, sheet) :
  ''' 企業がエクセルに既に書き込まれているかどうかのチェック
  '''

  gyouNo = 2

  result = False
  # 書き込む行番号の決定。空白行まで移動する
  while True :
    if sheet.cell(column = 2, row = gyouNo).value == None:
      result = False
      break
    if sheet.cell(column = 2, row = gyouNo).value == companyName:
      result = True
      break

    gyouNo += 1
  return result

def getValue(key, source):
  if key in source:
    return source[key]
  else :
    return ''

def witeExcel(list):
  # ファイルパスを指定
  # path = r'C:\Users\N030\Desktop\faceRecognition-master\faceRecognition\scraping'
  filename = '店舗情報.xlsx'
  
  # Bookの読み込み
  wb = openpyxl.load_workbook(filename)

  # シートの読み込み
  sheet = wb['店舗情報']
  
  # 既に会社情報が載っている場合は何もしない
  if kaisyaExist(list['companyName'], sheet) :
    return

  gyouNo = 1

  # 書き込む行番号の決定。空白行まで移動する
  while True :
    gyouNo += 1

    if sheet.cell(column = 1,row = gyouNo).value == None:
      sheet.cell(column = 1,row = gyouNo).value = gyouNo - 1
      break

  # 会社名
  sheet.cell(column = 2, row = gyouNo).value = list['companyName']

  # メールアドレス
  sheet.cell(column = 3, row = gyouNo).value = getValue('mail', list)

  # 郵便番号
  sheet.cell(column = 4, row = gyouNo).value = list['postCode']

  # 住所1
  sheet.cell(column = 5, row = gyouNo).value = list['adress1']

  # 住所2
  sheet.cell(column = 6, row = gyouNo).value = getValue('adress2', list)

  # 代表者
  sheet.cell(column = 7, row = gyouNo).value = getValue('representative', list)

  # ストア名
  sheet.cell(column = 8, row = gyouNo).value = getValue('shopName', list)

  # ストア名（フリガナ）
  sheet.cell(column = 9, row = gyouNo).value = getValue('shopNameKana', list)

  # ストア紹介
  sheet.cell(column = 10, row = gyouNo).value = getValue('setumei', list)

  # 運営責任者
  sheet.cell(column = 11, row = gyouNo).value = getValue('operationManager', list)

  # 電話番号
  sheet.cell(column = 12, row = gyouNo).value = getValue('tel', list)

  # お問い合わせファックス番号
  sheet.cell(column = 13, row = gyouNo).value = getValue('fax', list)
    
  # ストア営業日/時間
  sheet.cell(column = 14, row = gyouNo).value = getValue('Time', list)

  # 関連ストア
  sheet.cell(column = 15, row = gyouNo).value = getValue('relatedStore', list)    

  # ここで保存
  wb.save(filename)


# 商品名称を指定する
# main('消臭剤')
# main('雪塩ちんすこう（ミニ） 12個入（2×6袋)')
main('雪塩ちんすこう')
# main('ちんすこう')
# main('紅芋タルト')
# main('ゴーヤ茶 ティーパック')
